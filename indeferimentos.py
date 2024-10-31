from openpyxl import load_workbook, Workbook

planilha = load_workbook("planilhasjan_a_jun.xlsx")

dados = planilha.active

# Dicionário para mapear CNPJs com ocorrências totais
cnpj_count = {}
# Dicionário para mapear CNPJs com ocorrências "Indeferida"
cnpj_indeferida_count = {}

# Iteração para mapear todos os CNPJs e contar ocorrências
for row in range(2, dados.max_row + 1):
    cnpj = dados.cell(row=row, column=4).value
    if cnpj not in cnpj_count:
        cnpj_count[cnpj] = 0
    cnpj_count[cnpj] += 1
    # Verifica se a situação é "Indeferida"
    situacao = dados.cell(row=row, column=13).value
    if situacao and situacao.strip().lower() == "indeferida":
        if cnpj not in cnpj_indeferida_count:
            cnpj_indeferida_count[cnpj] = 0
        cnpj_indeferida_count[cnpj] += 1

print("CNPJS Mapeados...")

# Criando nova planilha
nova_planilha = Workbook()
indice = nova_planilha.active
indice.title = "Índice"

# Definindo Cabeçalhos

indice["A1"] = "CNPJ"
indice["B1"] = "Indeferidas"
indice["C1"] = "Total Analisadas"
indice["D1"] = "Índice de Indeferimento (%)"

# Inserindo dados na planilha
for linha, cnpj in enumerate(cnpj_indeferida_count, start=2):
    indice[f"A{linha}"] = cnpj
    indice[f"B{linha}"] = cnpj_indeferida_count[cnpj]
    indice[f"C{linha}"] = cnpj_count.get(cnpj, 0)

    # Calculando o indice de indeferimento
    total_analisadas = cnpj_count.get(cnpj, 0)
    indeferidas = cnpj_indeferida_count[cnpj]
    if total_analisadas > 0:
        indice_indeferimento = (indeferidas / total_analisadas) * 100
    else:
        indice_indeferimento = 0

    indice[f"D{linha}"] = round(indice_indeferimento, 2)

# Salvando a Planilha
nova_planilha.save(filename="planilhas/indeferimentojan - jun 2024.xlsx")

print("Planilha criada com sucesso")
