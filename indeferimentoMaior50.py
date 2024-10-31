from openpyxl import load_workbook

planilha = load_workbook('planilhas/pafal/indeferimento jan - jun 2024.xlsx')

dados = planilha.active

linhasExcluir = []

for indice in range(2, dados.max_row +1):
    if dados.cell(row=indice, column=4).value < 50:
        linhasExcluir.append(indice)
        
for indices in reversed(linhasExcluir):
    dados.delete_rows(indices)
    
planilha.save('planilhas/pafal/indeferimentos maior50 jan - jun 2024.xlsx')

print("finalizado")